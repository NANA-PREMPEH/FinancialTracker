import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_pdf(expenses, user, title='Transaction Report'):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, spaceAfter=12)
    elements.append(Paragraph(title, title_style))

    # Subtitle with user and date info
    subtitle = f"Generated for {user.name} on {datetime.utcnow().strftime('%d %B %Y')}"
    elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))

    if not expenses:
        elements.append(Paragraph("No transactions found.", styles['Normal']))
        doc.build(elements)
        buffer.seek(0)
        return buffer

    # Summary
    total_income = sum(e.amount for e in expenses if e.transaction_type == 'income')
    total_expense = sum(e.amount for e in expenses if e.transaction_type == 'expense')
    summary_data = [
        ['Total Income', f'{user.default_currency} {total_income:,.2f}'],
        ['Total Expenses', f'{user.default_currency} {total_expense:,.2f}'],
        ['Net', f'{user.default_currency} {total_income - total_expense:,.2f}'],
    ]
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.green),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.red),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))

    # Transaction table
    headers = ['Date', 'Description', 'Category', 'Type', 'Amount']
    data = [headers]

    for expense in expenses:
        data.append([
            expense.date.strftime('%d/%m/%Y') if expense.date else '',
            (expense.description or '')[:40],
            expense.category.name if expense.category else '',
            expense.transaction_type.title() if expense.transaction_type else '',
            f'{expense.amount:,.2f}',
        ])

    col_widths = [1*inch, 2.2*inch, 1.2*inch, 0.8*inch, 1.2*inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 0.3*inch))
    footer_text = f"FinTracker - {len(expenses)} transaction(s)"
    elements.append(Paragraph(footer_text, styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel(expenses, user, title='Transactions'):
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    income_font = Font(color='22C55E')
    expense_font = Font(color='EF4444')

    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = f'{title} - {user.name} - {datetime.utcnow().strftime("%d %B %Y")}'
    ws['A1'].font = Font(bold=True, size=14)

    # Headers
    headers = ['Date', 'Description', 'Category', 'Wallet', 'Type', 'Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_num, expense in enumerate(expenses, 4):
        ws.cell(row=row_num, column=1).value = expense.date.strftime('%d/%m/%Y') if expense.date else ''
        ws.cell(row=row_num, column=2).value = expense.description or ''
        ws.cell(row=row_num, column=3).value = expense.category.name if expense.category else ''
        ws.cell(row=row_num, column=4).value = expense.wallet.name if expense.wallet else ''
        ws.cell(row=row_num, column=5).value = (expense.transaction_type or '').title()

        amount_cell = ws.cell(row=row_num, column=6)
        amount_cell.value = expense.amount
        amount_cell.number_format = '#,##0.00'
        if expense.transaction_type == 'income':
            amount_cell.font = income_font
        elif expense.transaction_type == 'expense':
            amount_cell.font = expense_font

        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border

    # Auto-size columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15

    # Summary sheet
    summary = wb.create_sheet('Summary')
    total_income = sum(e.amount for e in expenses if e.transaction_type == 'income')
    total_expense = sum(e.amount for e in expenses if e.transaction_type == 'expense')

    summary['A1'] = 'Financial Summary'
    summary['A1'].font = Font(bold=True, size=14)
    summary.merge_cells('A1:B1')

    summary['A3'] = 'Total Income'
    summary['B3'] = total_income
    summary['B3'].number_format = '#,##0.00'
    summary['B3'].font = income_font

    summary['A4'] = 'Total Expenses'
    summary['B4'] = total_expense
    summary['B4'].number_format = '#,##0.00'
    summary['B4'].font = expense_font

    summary['A5'] = 'Net Balance'
    summary['B5'] = total_income - total_expense
    summary['B5'].number_format = '#,##0.00'
    summary['B5'].font = Font(bold=True, size=12)

    summary['A7'] = 'Total Transactions'
    summary['B7'] = len(expenses)

    summary.column_dimensions['A'].width = 20
    summary.column_dimensions['B'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
